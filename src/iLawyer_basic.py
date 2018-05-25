#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""ILawyer basic functions"""

import os
import sys
reload(sys)
sys.setdefaultencoding('utf8')
import openpyxl
from googletrans import Translator
import json

PATH_FOLDER_OF_FILE = os.path.dirname(os.path.abspath(__file__))

# bien backslash de thay doi duong dan link tren Mac/ Window
# duong dan tren Mac: /; tren Window: \
BACKSLASH = '\\'    # gia tri default: \
if '/' in PATH_FOLDER_OF_FILE:
    BACKSLASH = '/'
# DONE
# Quy uoc id: 101: Luat doanh nghiep, 102: Luat dat dai, 103: Luat Thue Tai san

# so luong row ung voi moi file dict001.xlsx, dict002.xlsx voi tung bo luat
NUM_ROWS_DICT = [[214, 372], [213, 300], [16, 99]]
# so luong row ung voi moi file train001.xlsx voi tung bo luat
NUM_ROWS_TRAIN = [[372], [300], [99]]

# file chua cac tu profanity
PROFANITY_TXT = 'profanity.txt'

def print_data_json(data, out_json):
    """print date (type dict) ra file json"""
    f = open(out_json, "w")
    json.dump(data, f, ensure_ascii=False)
    f.close()
# DONE


def get_data_from_json(inp_json):
    """get data from file json"""
    value = {}
    with open(inp_json) as json_data:
        value = json.load(json_data)
    return value
# DONE


# ###########: import thong tin dict tu file .py ra file json
# import tree_default
# value = tree_default.Tree
# print_data_json(value, '100.json')
# # Luat doanh nghiep
# import tree_101
# value = tree_101.Node
# print_data_json(value, '101.json')
# # Luat dat dai
# import tree_102
# value = tree_102.Node
# print_data_json(value, '102.json')
# # Luat thue tai san
# import tree_103
# value = tree_103.Node
# print_data_json(value, '103.json')

# import file json chua thong tin decision tree tuong ung voi cac dieu luat
LAW_TREE = []
LAW_TREE.append(get_data_from_json('100.json'))
LAW_TREE.append(get_data_from_json('101.json'))
LAW_TREE.append(get_data_from_json('102.json'))
LAW_TREE.append(get_data_from_json('103.json'))
# neu them dieu luat moi thi them mot dong, chang han nhu: LAW_TREE.append(get_data_from_json('104.json'))


def get_path_law_foler(law_id):
    """lay duong dan folder chua luat co id = law_id"""
    return PATH_FOLDER_OF_FILE.replace('src', 'data' + BACKSLASH + str(law_id))
# DONE


def get_path_file_name(law_id, file_name):
    """lay duong dan den file (file_name) cua luat co id = law_id"""
    return os.path.join(get_path_law_foler(law_id), file_name)
# DONE


def get_path_dict_txt_xlsx(law_id):
    """lay duong dan den file law_dict, normal_dict cua luat co id = law_id"""
    basic_dict_txt = os.path.join(get_path_law_foler(law_id), 'basic_dict.txt')
    basic_dict_xlsx = os.path.join(get_path_law_foler(law_id), 'basic_dict.xlsx')
    standard_dict_txt = os.path.join(get_path_law_foler(law_id), 'standard_dict.txt')
    standard_dict_xlsx = os.path.join(get_path_law_foler(law_id), 'standard_dict.xlsx')
    return basic_dict_txt, basic_dict_xlsx, standard_dict_txt, standard_dict_xlsx
# DONE





def process_delete_file(inp_file_name):
    """Delete file has a name inp_file_name"""
    cmd = "rm -rf " + inp_file_name
    os.system(cmd)
    return 0
# DONE


def gen_column_from_xlsx(inp_xlsx, num_rows, id_column):
    """Read data in id_column col from file inp_xlsx from the second row"""
    data = []
    book = openpyxl.load_workbook(inp_xlsx)
    sheet = book.active
    for row in range(2, num_rows + 1): # ignore 1st row which is preserved for column title
        datum = sheet.cell(row=row, column=id_column).value
        data.append(datum)
    return data
# DONE


def get_article_from_prediction(inp_xlsx, article_number, id_column):
    """get article has id = article_number from file inp_xlsx, colums: id_columns """
    book = openpyxl.load_workbook(inp_xlsx)
    sheet = book.active
    article = sheet.cell(row=article_number + 1, column=id_column).value
    # remove one character '\n'
    replace = chr(10) + chr(10)
    article = article.replace(replace, "\n")
    return article
# DONE


def print_txt_from_string(string_value, out_txt):
    """print string_value (string_value either str type or list-str type) to txt file"""

    f = open(out_txt, "w")
    # check type(string_value) is str or list
    if isinstance(string_value, (list,)):
        for each_string in string_value:
            f.write(str(each_string) + '\n')
    else:
        f.write(str(string_value))

    f.close()
    return 0
# DONE


def process_vncore(inp_txt, out_txt):
    """Word segmentation from file: inp_txt, output file: out_txt"""
    cmd = "java RDRsegmenter " + inp_txt + " " + out_txt
    os.system(cmd)
    return 0
# DONE


def gen_array_from_txt(inp_txt):
    """generate string array from txt file"""
    string_array = []
    f = open(inp_txt, 'r')
    for each_string in f:
        str = each_string
        if ord(str[len(str) - 1]) == 10:    # check str has a break line
            str = str[:-1]
        string_array.append(str)
    f.close()
    return string_array
# DONE


def gen_distinct_array(inp_arr):
    """remove identical element from an array"""
    out_arr = []
    dicts = {}
    for element in inp_arr:
        value = dicts.get(element, 0)
        dicts[element] = value + 1
    for key, value in dicts.iteritems():
        out_arr.append(key)
    return out_arr
# DONE


def gen_words_from_vncore_out_txt(vn_core_out_txt):
    """
    # generate words (string array) from output of vncore
    # manipulation includes: ignore all words starting with digit,
    # replace underscore by space and change all characters to lower case
    """
    words_replicated = []
    lines = gen_array_from_txt(vn_core_out_txt)
    for line in lines:
        text = line
        if not text[0].isdigit():   # ignore all words starting with digit
            # replace underscore by space, change all characters to lower case
            word = text.replace('_', ' ').lower()
            words_replicated.append(word)
    # replicated words will be handled in gen_distinct_array
    return words_replicated
# DONE


def gen_feature_vector(string_array, standard_dict_txt):
    """generate feature vector (numeric array)
    standard_dict_txt: file chua tu dien normal
    """

    dict_array = gen_array_from_txt(standard_dict_txt)

    # remove all the string duplicate in string_array by dictinary - STL python
    dict_string_array = {}

    for element in string_array:
        value = dict_string_array.get(element, 0)
        dict_string_array[element] = value + 1

    feature_vector = []
    for word in dict_array:
        if word in dict_string_array:
            feature_vector.append(1)
        else:
            feature_vector.append(0)

    return feature_vector
# DONE


def gen_string_array_from_text(text):
    """generate string array from text, text either str type or list (str) type"""
    print_txt_from_string(text, 'vncore_inp.txt')
    process_vncore('vncore_inp.txt', 'vncore_out.txt')

    string_array = gen_words_from_vncore_out_txt('vncore_out.txt')

    process_delete_file('vncore_inp.txt')
    process_delete_file('vncore_out.txt')

    return string_array
# DONE


def gen_input_vector_from_text(text, standard_dict_txt):
    """generate feature vector (numeric 1D array) from a question (text)"""
    string_array = gen_string_array_from_text(text)
    inp_vector = gen_feature_vector(string_array, standard_dict_txt)
    return inp_vector
# DONE


def gen_feature_table_from_xlsx(inp_xlsx, num_rows, id_column_question, standard_dict_txt):
    "generate feature table (numeric 2D array) from a training set stored in xlsx"
    separator = 'separator'
    feature_table = []
    all_questions = gen_column_from_xlsx(inp_xlsx, num_rows, id_column_question)

    questions_with_seperator = []
    for question in all_questions:
        # add 'separator' before each question for convenient process
        questions_with_seperator.append(separator)
        questions_with_seperator.append(question)

    print_txt_from_string(questions_with_seperator, 'vncore_inp.txt')
    process_vncore('vncore_inp.txt', 'vncore_out.txt')
    string_array = gen_words_from_vncore_out_txt('vncore_out.txt')

    # separating question and get its feature vector
    question = []
    for each_string in string_array:
        if each_string == separator:
            if not question:
                continue
            feature_vector = gen_feature_vector(question, standard_dict_txt)
            feature_table.append(feature_vector)
            question = []
        else:
            question.append(each_string)

    # add a last feature_vector question
    feature_vector = gen_feature_vector(question, standard_dict_txt)
    feature_table.append(feature_vector)

    process_delete_file('vncore_inp.txt')
    process_delete_file('vncore_out.txt')

    return feature_table
# DONE


def trans_string_into_other_lang(origin_string, dest_lang):
    """dich mot string sang ngon ngu dest_lang"""
    translator = Translator()
    # translator.detect(origin_string)  # xac dinh ngon ngu
    return translator.translate(origin_string, dest=dest_lang).text
# DONE


def check_string_array_has_at_least_2_keywords(string_array, basic_dict_txt):
    """kiem tra string_array (string_array thu duoc bang viec tach tu)
    co it nhat 2 key_word trong bo luat tuong ung voi basic_dict_txt"""
    keywords = gen_array_from_txt(basic_dict_txt)
    # counting number of string_array list in keywords
    cnt = 0
    for word in string_array:
        if word in keywords:
            cnt += 1
    return cnt > 1
# DONE


def check_string_array_has_profanity(string_array, profanity_txt):
    """kiem tra string_array (string_array thu duoc bang viec tach tu)
    co chua profanity khong, profanity duoc load tu file profanity_txt"""
    keywords = gen_array_from_txt(profanity_txt)
    # counting number of string_array list in keywords
    cnt = 0
    for word in string_array:
        if word in keywords:
            return True
    return False
# DONE



def get_article(type_law, num_article, lang, id_column):
    """lay dieu luat num_artilce trong file inp_xlsx o cot 2 cua file .xlsx"""
    law_xlsx = str(type_law) + '.xlsx'
    answer = get_article_from_prediction(law_xlsx, num_article, id_column=id_column)
    # answer = trans_string_into_other_lang(answer, lang)
    array_answer = []
    len_answer = len(answer)
    # do gioi han moi tin nhan tren messenger là 2000 ki tu
    # do vay tach dieu luat thanh mot array, moi phan tu cua array la mot string co do dai 1500 ki tu
    for i in range(0, (len_answer + 1499) / 1500):
        value = answer[(i * 1500): ((i + 1) * 1500)]
        array_answer.append(value)
    return array_answer
# DONE


# tra ve payload tuong ung voi title = message trong quick_reply_array
# neu ko ton tai title = message thi gia tri tra ve = -1
def get_payload_from_text(message, quick_reply_array):
    message = message.lower()
    for dict in quick_reply_array:
        title = dict['title'].lower()
        pay_load = dict['payload']
        if message == title:
            return pay_load
    return -1










