#!/usr/bin/env python
# -*- coding: utf-8 -*-

import sys
reload(sys)
sys.setdefaultencoding('utf8')
from googletrans import Translator
import tree_default as td
import openpyxl
import json
import iLawyer_basic as ib
from xlrd import open_workbook
import xlrd
#
def trans_string_into_ja(origin_string, dest_lang):
    """dich mot string sang ngon ngu dest_lang"""
    translator = Translator()
    # translator.detect(origin_string)  # xac dinh ngon ngu
    return translator.translate(origin_string, dest=dest_lang).text
# DONE

# dich mot array string sang tieng nhat va viet ra file txt
def trans_array_string_into_ja(array_string, out_txt):
    translator = Translator()
    translations = translator.translate(array_string, dest='ja')

    f = open(out_txt, 'w')
    for trans in translations:
        f.write(trans.origin + '\n')
        f.write(trans.text + '\n\n')
    f.close()
# DONE


# dich mot array string sang tieng nhat va viet ra file txt
def trans_file_into_vi(inp_txt, out_txt):
    translator = Translator()
    array_string = []
    fr = open(inp_txt, 'r')
    for word in fr:
        print word[:-1]
        array_string.append(word[:-1])

    translations = translator.translate(array_string, dest='vi')

    fw = open(out_txt, 'w')
    for trans in translations:
        fw.write(trans.origin + '\n')
        fw.write(trans.text + '\n\n')
    fw.close()

# DONE

#
# print len(ib.LAW_TREE[0])
# # for k, v in ib.LAW_TREE[0].iteritems():
# #     print k
# print ib.LAW_TREE[0]['0']['text']
# print ib.LAW_TREE[0]['0']['quick_reply']
# print ib.LAW_TREE[0]['profanity_ans']['text']['vi']

# book = openpyxl.load_workbook('test101.xlsx')
# sheet = book.active
# for row in range(2, 2 + 1):  # ignore 1st row which is preserved for column title
#     datum1 = sheet.cell(row=row, column=1).value
#     datum2 = sheet.cell(row=row, column=1).value
#     # print datum1.encode('utf-8'), "\nTuan\n", datum2.encode('utf-8')
#     print type(datum1)
#     print len(datum1)
#     for i in range(0, 5):
#         print datum1[i].encode('utf-8'), " ", datum2[i].encode('utf-8')

    # data.append(datum)
# return data
#
# book = xlrd.open_workbook("test101.xlsx")
# print("The number of worksheets is {0}".format(book.nsheets))
# print("Worksheet name(s): {0}".format(book.sheet_names()))
# sh = book.sheet_by_index(0)
# print("{0} {1} {2}".format(sh.name, sh.nrows, sh.ncols))
#
# print("Cell A1 is {0}".format(sh.cell_value(rowx=1, colx=0)))
# for rx in range(sh.nrows):
#     print(sh.row(rx))
# for rowx in range(sh.nrows):
#     for colx in range(sh.ncols):
#         cty = sh.cell_type(rowx, colx)
#         print rowx, " ", colx, " ", cty

# fmt = wb.xf_list[cell.xf_index]
# print("type(fmt) is", type(fmt))
# print("Dumped Info:")
# fmt.dump()
# path = 'test101.xlsx'
# wb = open_workbook(path, formatting_info=True)
# sheet = wb.sheet_by_name("Sheet1")
# cell = sheet.cell(0, 0) # The first cell
# print("cell.xf_index is", cell.xf_index)
# fmt = wb.xf_list[cell.xf_index]
# print("type(fmt) is", type(fmt))
# print("Dumped Info:")
# fmt.dump()

#
# # START CRAWLER DATA
# import urllib
# from bs4 import BeautifulSoup
# import requests
#
# URL_TAX = 'https://thuvienphapluat.vn/van-ban/Thue-Phi-Le-Phi/Luat-thue-tai-san-379974.aspx'
# r = requests.get(URL_TAX)
# data = r.text
# soup = BeautifulSoup(data, 'html.parser')
# print "LEN: ", len(soup.find_all('p', attrs={'style': "margin-top:6.0pt"}))
#
# fw = open('103.txt', 'w')
#
# for company in soup.find_all('p', attrs={'style': "margin-top:6.0pt"}):
#     # print company, "\n\n"
#     fw.write(str(company) + '\n\n')
# # print soup